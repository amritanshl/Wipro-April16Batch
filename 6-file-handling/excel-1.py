from openpyxl import Workbook, load_workbook
data = [
    [
        "Account_No","Customer_Name","Balance","City"
    ],
    [
        107,"Akshat Lal", 484908.99, "New Delhi"
    ],
    [
        108,"Harshita", 7674.99, "Hyderabad"
    ]
]

wb = load_workbook('customers.xlsx')
ws = wb.active
ws.title = "Customers"

for row in data:
    pass
#     ws.append(row)
# wb.save('customers.xlsx')

# for row in ws.iter_rows(values_only=True):
#     print(row)

for row in range(2, ws.max_row + 1):
    print(ws.cell(row=row, column=2).value)
